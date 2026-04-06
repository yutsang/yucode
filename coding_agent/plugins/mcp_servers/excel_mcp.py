"""MCP server for Excel (.xlsx) file operations.

Requires: openpyxl >= 3.1

Run:  python -m coding_agent.plugins.mcp_servers.excel_mcp
"""

from __future__ import annotations

import json
from typing import Any

from ._protocol import McpStdioServer


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl>=3.1")


def handle_list_sheets(args: dict[str, Any]) -> Any:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(args["path"], read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return {"sheets": sheets}


def handle_read_sheet(args: dict[str, Any]) -> Any:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(args["path"], read_only=True, data_only=True)
    sheet_name = args.get("sheet")
    ws = wb[sheet_name] if sheet_name else wb.active
    max_rows = int(args.get("max_rows", 500))

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()

    if not rows:
        return {"headers": [], "rows": []}
    return {"headers": rows[0], "rows": rows[1:]}


def handle_read_cell(args: dict[str, Any]) -> Any:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(args["path"], read_only=True, data_only=True)
    sheet_name = args.get("sheet")
    ws = wb[sheet_name] if sheet_name else wb.active
    cell = args["cell"]
    value = ws[cell].value
    wb.close()
    return {"cell": cell, "value": str(value) if value is not None else ""}


def handle_write_cell(args: dict[str, Any]) -> Any:
    openpyxl = _require_openpyxl()
    path = args["path"]
    wb = openpyxl.load_workbook(path)
    sheet_name = args.get("sheet")
    ws = wb[sheet_name] if sheet_name else wb.active
    cell = args["cell"]
    ws[cell] = args["value"]
    wb.save(path)
    wb.close()
    return {"written": cell, "value": args["value"]}


def handle_sheet_to_json(args: dict[str, Any]) -> Any:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(args["path"], read_only=True, data_only=True)
    sheet_name = args.get("sheet")
    ws = wb[sheet_name] if sheet_name else wb.active
    max_rows = int(args.get("max_rows", 500))

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:max_rows + 1]:
        record = {}
        for j, val in enumerate(row):
            key = headers[j] if j < len(headers) else f"col_{j}"
            record[key] = str(val) if val is not None else ""
        result.append(record)
    return result


def main():
    server = McpStdioServer("yucode-excel")

    server.register_tool("list_sheets", "List sheet names in an Excel file.", {
        "type": "object",
        "properties": {"path": {"type": "string", "description": "Path to .xlsx file"}},
        "required": ["path"],
    }, handle_list_sheets)

    server.register_tool("read_sheet", "Read an Excel sheet as a table (headers + rows).", {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Sheet name (default: active sheet)"},
            "max_rows": {"type": "integer", "description": "Maximum rows to read (default 500)"},
        },
        "required": ["path"],
    }, handle_read_sheet)

    server.register_tool("read_cell", "Read a single cell value.", {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "cell": {"type": "string", "description": "Cell reference like A1, B2"},
        },
        "required": ["path", "cell"],
    }, handle_read_cell)

    server.register_tool("write_cell", "Write a value to a single cell.", {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "cell": {"type": "string"},
            "value": {"type": "string"},
        },
        "required": ["path", "cell", "value"],
    }, handle_write_cell)

    server.register_tool("sheet_to_json", "Convert a sheet to a list of JSON records.", {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string"},
            "max_rows": {"type": "integer"},
        },
        "required": ["path"],
    }, handle_sheet_to_json)

    server.serve()


if __name__ == "__main__":
    main()
